import re
import requests
import matplotlib.pyplot as plt
import pandas as pd
import numpy as np
import statistics
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.utils.dataframe import dataframe_to_rows
from io import BytesIO

API_KEY = "8672905b631a8a0b3a41a62affffec7f"
RE_FECHA = re.compile(r"^\d{4}-\d{2}-\d{2}$")

diccionario_curioso = {
    "accion": [1, 2, 3],
    "comedia": [4, 5, 6],
    "drama": [7, 8, 9]
}

class TMDbAPI:
    def __init__(self, api_key):
        self.api_key = api_key
        self.base_url = "https://api.themoviedb.org/3"

    def obtener_generos(self):
        url = f"{self.base_url}/genre/movie/list?api_key={self.api_key}&language=es"
        return requests.get(url).json()["genres"]

    def buscar_peliculas(self, genero, desde, hasta):
        url = (
            f"{self.base_url}/discover/movie?api_key={self.api_key}&language=es"
            f"&sort_by=vote_average.desc&vote_count.gte=100"
            f"&with_genres={genero}&primary_release_date.gte={desde}&primary_release_date.lte={hasta}&page=1"
        )
        return requests.get(url).json()["results"]

    def obtener_peliculas(self, generos, desde, hasta, top_n):
        peliculas_totales = []
        for genero in generos:
            peliculas_totales.extend(self.buscar_peliculas(genero, desde, hasta))
        peliculas_totales = sorted(
            {p['id']: p for p in peliculas_totales}.values(),
            key=lambda x: x['vote_average'], reverse=True
        )[:top_n]
        return peliculas_totales

def obtener_datos_peliculas():
    api = TMDbAPI(API_KEY)
    generos_disponibles = api.obtener_generos()

    print("\nGÉNEROS DISPONIBLES:")
    for i, genero in enumerate(generos_disponibles, start=1):
        print(f"{i}. {genero['name']}")
    print(f"{len(generos_disponibles)+1}. Todos los géneros")

    while True:
        try:
            cantidad_generos = int(input("\n¿Cuántos géneros deseas seleccionar? "))
            if 1 <= cantidad_generos <= len(generos_disponibles) + 1:
                break
        except ValueError:
            continue

    if cantidad_generos == len(generos_disponibles) + 1:
        generos_seleccionados = [g["id"] for g in generos_disponibles]
    else:
        generos_seleccionados = []
        for i in range(cantidad_generos):
            while True:
                try:
                    seleccion = int(input(f"Selecciona el género #{i+1} (1-{len(generos_disponibles)}): "))
                    if 1 <= seleccion <= len(generos_disponibles):
                        generos_seleccionados.append(generos_disponibles[seleccion - 1]["id"])
                        break
                except ValueError:
                    continue

    while True:
        fecha_inicio = input("Fecha de inicio (YYYY-MM-DD): ")
        if RE_FECHA.match(fecha_inicio): break

    while True:
        fecha_fin = input("Fecha de fin (YYYY-MM-DD): ")
        if RE_FECHA.match(fecha_fin): break

    while True:
        try:
            top_n = int(input("¿Cuántas películas deseas mostrar? "))
            if top_n > 0: break
        except ValueError:
            continue

    peliculas = api.obtener_peliculas(generos_seleccionados, fecha_inicio, fecha_fin, top_n)
    return peliculas

def generar_graficas(peliculas):
    df = pd.DataFrame(peliculas)[['title', 'vote_average', 'release_date']]

    plt.figure()
    df.sort_values('vote_average').plot.barh(x='title', y='vote_average', legend=False, color='skyblue')
    plt.title("Gráfico de Barras - Puntuaciones")
    plt.xlabel("Puntuación")
    plt.grid(True)
    plt.tight_layout()
    plt.savefig("grafico_barras.png")
    plt.close()

    plt.figure()
    plt.plot(df['title'], df['vote_average'], marker='o', color='orange')
    plt.xticks(rotation=90)
    plt.title("Gráfico de Líneas - Puntuaciones")
    plt.ylabel("Puntuación")
    plt.grid(True)
    plt.tight_layout()
    plt.savefig("grafico_lineas.png")
    plt.close()

    plt.figure()
    plt.scatter(range(len(df)), df['vote_average'], c='red')
    plt.title("Diagrama de Dispersión")
    plt.xlabel("Película (índice)")
    plt.ylabel("Puntuación")
    plt.grid(True)
    plt.tight_layout()
    plt.savefig("grafico_dispersion.png")
    plt.close()

    plt.figure()
    top5 = df.nlargest(5, 'vote_average')
    plt.pie(top5['vote_average'], labels=top5['title'], autopct='%1.1f%%')
    plt.title("Top 5 Puntuaciones - Gráfico de Pastel")
    plt.tight_layout()
    plt.savefig("grafico_pastel.png")
    plt.close()

def guardar_en_excel(peliculas):
    df = pd.DataFrame(peliculas)[['title', 'vote_average', 'release_date']]
    wb = Workbook()
    ws_data = wb.active
    ws_data.title = "Datos"

    for r in dataframe_to_rows(df, index=False, header=True):
        ws_data.append(r)

    ws_metricas = wb.create_sheet("Métricas")
    puntuaciones = df['vote_average']

    media = np.mean(puntuaciones)
    mediana = np.median(puntuaciones)
    moda = statistics.mode(puntuaciones)
    desviacion = np.std(puntuaciones)

    ws_metricas.append(["Métrica", "Valor"])
    ws_metricas.append(["Media", media])
    ws_metricas.append(["Mediana", mediana])
    ws_metricas.append(["Moda", moda])
    ws_metricas.append(["Desviación estándar", desviacion])

    ws_graficas = wb.create_sheet("Gráficas")
    for nombre, celda in zip(["grafico_barras.png", "grafico_lineas.png", "grafico_dispersion.png", "grafico_pastel.png"], ["A1", "A30", "A60", "A90"]):
        with open(nombre, 'rb') as f:
            img_data = BytesIO(f.read())
        img = ExcelImage(img_data)
        ws_graficas.add_image(img, celda)

    wb.save("peliculas_analisis.xlsx")
